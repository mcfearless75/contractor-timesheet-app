from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    login_required,
    logout_user,
    current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy.exc import OperationalError
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

app = Flask(__name__)
app.config['SECRET_KEY'] = 'supersecretkey'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timesheets.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ------------------------ Models ------------------------

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    security_answer = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(20), default='contractor')  # 'contractor' or 'manager'

class Timesheet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contractor_name = db.Column(db.String(100))
    client = db.Column(db.String(100))
    site_address = db.Column(db.String(200))
    week_start = db.Column(db.Date)
    week_end = db.Column(db.Date)
    basic_hours = db.Column(db.Float)
    saturday_hours = db.Column(db.Float)
    sunday_hours = db.Column(db.Float)
    hourly_rate = db.Column(db.Float)
    total_hours = db.Column(db.Float)
    calculated_pay = db.Column(db.Float)
    approved = db.Column(db.Boolean, default=False)
    submitted_on = db.Column(db.DateTime, default=datetime.utcnow)
    approved_on = db.Column(db.DateTime, nullable=True)

# ---------------------- User Loader ----------------------

@login_manager.user_loader
def load_user(user_id):
    try:
        # Attempt to load the user normally
        return User.query.get(int(user_id))
    except OperationalError:
        # If schema is out-of-date (e.g. missing security_answer), rebuild DB
        db.drop_all()
        db.create_all()
        return None

# ------------------------ Routes ------------------------

# ----------------------------------------------------------
# home
# ----------------------------------------------------------
@app.route('/')
def home():
    return render_template('index.html')

# ----------------------------------------------------------
# register
# ----------------------------------------------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        email = request.form['email'].strip().lower()
        password = request.form['password']
        answer = request.form['security_answer'].strip()

        if not username or not email or not password or not answer:
            flash('All fields are required.', 'danger')
            return redirect(url_for('register'))

        existing_user = User.query.filter(
            (User.username == username) | (User.email == email)
        ).first()
        if existing_user:
            flash('Username or email already exists.', 'warning')
            return redirect(url_for('register'))

        hashed_pw = generate_password_hash(password)
        new_user = User(
            username=username,
            email=email,
            password=hashed_pw,
            security_answer=answer,
            role='contractor'  # default role
        )
        db.session.add(new_user)
        db.session.commit()

        flash('Account created successfully! Please log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

# ----------------------------------------------------------
# login
# ----------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']

        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            login_user(user)
            # Redirect contractors to submit page, managers to dashboard
            if user.role == 'manager':
                return redirect(url_for('dashboard'))
            else:
                return redirect(url_for('submit_timesheet'))

        flash('Invalid credentials. Please try again.', 'danger')
    return render_template('login.html')

# ----------------------------------------------------------
# reset password (security question)
# ----------------------------------------------------------
@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        username = request.form['username'].strip()
        answer = request.form['security_answer'].strip()
        new_pw = request.form['new_password']

        user = User.query.filter_by(username=username).first()
        if user and user.security_answer.strip().lower() == answer.lower():
            user.password = generate_password_hash(new_pw)
            db.session.commit()
            flash('Password successfully reset. Please log in.', 'success')
            return redirect(url_for('login'))

        flash('Incorrect security answer or username not found.', 'danger')

    return render_template('reset_password.html')

# ----------------------------------------------------------
# logout
# ----------------------------------------------------------
@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('home'))

# ----------------------------------------------------------
# submit timesheet (contractor only)
# ----------------------------------------------------------
@app.route('/submit', methods=['GET', 'POST'])
@login_required
def submit_timesheet():
    if current_user.role != 'contractor':
        flash('Only contractors can submit timesheets.', 'warning')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        contractor_name = current_user.username
        client = request.form['client']
        site_address = request.form['site_address']
        week_start = datetime.strptime(request.form['week_start'], '%Y-%m-%d').date()
        week_end = datetime.strptime(request.form['week_end'], '%Y-%m-%d').date()
        basic_hours = float(request.form['basic_hours'])
        saturday_hours = float(request.form['saturday_hours'])
        sunday_hours = float(request.form['sunday_hours'])
        hourly_rate = float(request.form['hourly_rate'])

        total_hours = basic_hours + saturday_hours + sunday_hours
        calculated_pay = (
            basic_hours * hourly_rate +
            saturday_hours * hourly_rate * 1.5 +
            sunday_hours * hourly_rate * 1.75
        )

        ts = Timesheet(
            contractor_name=contractor_name,
            client=client,
            site_address=site_address,
            week_start=week_start,
            week_end=week_end,
            basic_hours=basic_hours,
            saturday_hours=saturday_hours,
            sunday_hours=sunday_hours,
            hourly_rate=hourly_rate,
            total_hours=total_hours,
            calculated_pay=calculated_pay
        )
        db.session.add(ts)
        db.session.commit()
        return redirect(url_for('thank_you'))

    return render_template('submit.html')

# ----------------------------------------------------------
# thank you page after submission
# ----------------------------------------------------------
@app.route('/thank_you')
@login_required
def thank_you():
    return render_template('thank_you.html')

# ----------------------------------------------------------
# manager dashboard (view & approve timesheets)
# ----------------------------------------------------------
@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role != 'manager':
        abort(403)

    timesheets = Timesheet.query.order_by(Timesheet.submitted_on.desc()).all()
    return render_template('dashboard.html', timesheets=timesheets)

# ----------------------------------------------------------
# approve a single timesheet
# ----------------------------------------------------------
@app.route('/approve/<int:ts_id>')
@login_required
def approve_timesheet(ts_id):
    if current_user.role != 'manager':
        abort(403)

    ts = Timesheet.query.get_or_404(ts_id)
    ts.approved = True
    ts.approved_on = datetime.utcnow()
    db.session.commit()
    flash('Timesheet approved.', 'success')
    return redirect(url_for('dashboard'))

# ----------------------------------------------------------
# export approved timesheets to Excel
# ----------------------------------------------------------
@app.route('/export')
@login_required
def export_timesheets():
    if current_user.role != 'manager':
        abort(403)

    timesheets = Timesheet.query.filter_by(approved=True).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Timesheets"

    headers = [
        "Contractor", "Client", "Site", "Week Start", "Week End",
        "Basic Hours", "Saturday Hours", "Sunday Hours",
        "Hourly Rate", "Total Hours", "Calculated Pay", "Submitted On"
    ]
    ws.append(headers)
    for col in ws[1]:
        col.font = Font(bold=True)
        col.alignment = Alignment(horizontal='center')

    for ts in timesheets:
        ws.append([
            ts.contractor_name,
            ts.client,
            ts.site_address,
            ts.week_start.strftime('%Y-%m-%d'),
            ts.week_end.strftime('%Y-%m-%d'),
            ts.basic_hours,
            ts.saturday_hours,
            ts.sunday_hours,
            ts.hourly_rate,
            ts.total_hours,
            ts.calculated_pay,
            ts.submitted_on.strftime('%Y-%m-%d %H:%M')
        ])

    for column in ws.columns:
        max_len = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name="Approved_Timesheets.xlsx", as_attachment=True)

# ----------------------------------------------------------
# one-time routes to manage DB (will be disabled after use)
# ----------------------------------------------------------

@app.route('/delete-db')
def delete_db():
    return "⛔️ Disabled for security reasons."

@app.route('/initdb')
def initdb():
    return "⛔️ Disabled for security reasons."

# ------------------------ Run ------------------------

if __name__ == '__main__':
    app.run(debug=True)
